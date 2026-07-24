"""
教师端 API 路由
"""
import math
import os
import uuid
import shutil
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import Optional, List

from ...core.database import get_db
from ...core.config import settings
from ...api.v1.deps import get_current_teacher, get_current_active_user
from ...models.user import User
from ...models.chapter import Chapter
from ...models.course_material import CourseMaterial
from ...models.exam_config import ExamConfig
from ...models.exam_record import ExamRecord
from ...models.teacher_student import TeacherStudent
from ...models.wrong_answer import WrongAnswer
from ...schemas.teacher import (
    MaterialResponse, StudentInfo, StudentExamResult,
    ModelConfigUpdate, ModelConfigResponse, ChapterResponse, DashboardStats
)
from ...schemas.exam import ExamConfigCreate, ExamConfigResponse
from ...services.rag.retriever import add_text as add_to_rag
from ...utils.file_parser import parse_file

router = APIRouter(prefix="/teacher", tags=["教师端"])

# ── 资料管理 ─────────────────────────────────────

@router.post("/materials/upload")
async def upload_material(
    file: UploadFile = File(...),
    chapter_id: Optional[int] = Form(None),
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """上传课件资料（PDF/PPTX/DOCX/MD/TXT）"""
    # 校验文件类型
    ext = os.path.splitext(file.filename)[1].lower()
    allowed = {'.pdf', '.pptx', '.ppt', '.docx', '.doc', '.md', '.txt', '.cpp', '.h', '.hpp'}
    if ext not in allowed:
        raise HTTPException(400, f"不支持的文件类型: {ext}")

    # 保存文件（使用 ASCII 安全文件名避免 Windows 编码问题）
    file_id = uuid.uuid4().hex[:12]
    safe_name = f"{file_id}{ext}"
    save_path = os.path.join(settings.UPLOAD_DIR, safe_name)

    content = await file.read()
    with open(save_path, 'wb') as f:
        f.write(content)

    # 尝试解析文本（失败不影响上传）
    text_content = ""
    try:
        text_content = parse_file(save_path)
    except Exception as e:
        print(f"[WARN] File parse skipped: {e}")

    # 查询章节标题
    chapter_title = ""
    if chapter_id:
        ch = db.query(Chapter).filter(Chapter.id == chapter_id).first()
        if ch:
            chapter_title = ch.title

    # 尝试存入 RAG 向量库（失败不影响上传）
    try:
        if text_content:
            add_to_rag(text_content, {
                "filename": file.filename,
                "file_type": ext,
                "chapter_id": str(chapter_id) if chapter_id else "",
                "chapter_title": chapter_title,
                "teacher_id": str(current_user.id),
            })
    except Exception as e:
        print(f"[WARN] RAG index skipped: {e}")

    # 保存数据库记录
    material = CourseMaterial(
        teacher_id=current_user.id,
        chapter_id=chapter_id,
        file_name=file.filename,
        file_path=safe_name,
        file_type=ext,
        parsed_content=text_content[:2000] if text_content else None,
    )
    db.add(material)
    db.commit()
    db.refresh(material)

    return {"id": material.id, "file_name": file.filename, "message": "上传成功"}


@router.get("/materials")
def list_materials(
    chapter_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """列出已上传的资料（分页）"""
    q = db.query(CourseMaterial).filter(CourseMaterial.teacher_id == current_user.id)
    if chapter_id is not None:
        q = q.filter(CourseMaterial.chapter_id == chapter_id)

    total = q.count()
    materials = q.order_by(CourseMaterial.uploaded_at.desc()).offset(
        (page - 1) * page_size).limit(page_size).all()

    result = []
    for m in materials:
        chapter_title = None
        if m.chapter_id:
            ch = db.query(Chapter).filter(Chapter.id == m.chapter_id).first()
            if ch:
                chapter_title = ch.title
        result.append(MaterialResponse(
            id=m.id,
            file_name=m.file_name,
            file_type=m.file_type,
            chapter_id=m.chapter_id,
            chapter_title=chapter_title,
            uploaded_at=m.uploaded_at.isoformat() if m.uploaded_at else "",
        ))
    return {
        "items": result,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@router.delete("/materials/{material_id}")
def delete_material(
    material_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """删除资料"""
    material = db.query(CourseMaterial).filter(
        CourseMaterial.id == material_id,
        CourseMaterial.teacher_id == current_user.id
    ).first()
    if not material:
        raise HTTPException(404, "资料不存在")

    # 删除文件
    file_path = os.path.join(settings.UPLOAD_DIR, material.file_path)
    if os.path.exists(file_path):
        os.remove(file_path)

    db.delete(material)
    db.commit()
    return {"message": "删除成功"}


@router.post("/materials/batch-delete")
def batch_delete_materials(
    data: dict,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """批量删除资料"""
    material_ids = data.get("material_ids", [])
    if not material_ids:
        raise HTTPException(400, "material_ids 不能为空")

    deleted = 0
    failed = 0
    errors = []

    for mid in material_ids:
        try:
            material = db.query(CourseMaterial).filter(
                CourseMaterial.id == mid,
                CourseMaterial.teacher_id == current_user.id
            ).first()

            if not material:
                failed += 1
                errors.append(f"资料 ID {mid}: 不存在或无权删除")
                continue

            # 删除文件
            file_path = os.path.join(settings.UPLOAD_DIR, material.file_path)
            if os.path.exists(file_path):
                os.remove(file_path)
            else:
                errors.append(f"资料 ID {mid}: 文件不存在于磁盘")

            db.delete(material)
            deleted += 1

        except Exception as e:
            failed += 1
            errors.append(f"资料 ID {mid}: {str(e)}")

    db.commit()

    return {"deleted": deleted, "failed": failed, "errors": errors}


# ── 考核管理 ─────────────────────────────────────

@router.post("/exams", response_model=ExamConfigResponse)
def create_exam(
    config: ExamConfigCreate,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """配置章节考核"""
    # 验证章节存在
    chapter = db.query(Chapter).filter(Chapter.id == config.chapter_id).first()
    if not chapter:
        raise HTTPException(404, "章节不存在")

    # 验证题型数量
    total_by_type = config.choice_count + config.truefalse_count + config.essay_count + config.programming_count
    if total_by_type != config.total_questions:
        raise HTTPException(400, f"题型数量之和({total_by_type})不等于总题数({config.total_questions})")

    exam = ExamConfig(
        chapter_id=config.chapter_id,
        teacher_id=current_user.id,
        total_questions=config.total_questions,
        choice_count=config.choice_count,
        truefalse_count=config.truefalse_count,
        essay_count=config.essay_count,
        programming_count=config.programming_count,
        time_limit_minutes=config.time_limit_minutes,
    )
    exam.set_knowledge_points(config.knowledge_points)
    exam.set_evaluation_dimensions(config.evaluation_dimensions)
    db.add(exam)
    db.commit()
    db.refresh(exam)

    return ExamConfigResponse(
        id=exam.id,
        chapter_id=exam.chapter_id,
        teacher_id=exam.teacher_id,
        total_questions=exam.total_questions,
        choice_count=exam.choice_count,
        truefalse_count=exam.truefalse_count,
        essay_count=exam.essay_count,
        programming_count=exam.programming_count or 0,
        knowledge_points=exam.get_knowledge_points(),
        evaluation_dimensions=exam.get_evaluation_dimensions(),
        created_at=exam.created_at.isoformat() if exam.created_at else "",
        chapter_title=chapter.title,
    )


@router.get("/exams")
def list_exams(
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """查看已配置的考核（分页）"""
    q = db.query(ExamConfig).filter(
        ExamConfig.teacher_id == current_user.id
    ).order_by(ExamConfig.created_at.desc())

    total = q.count()
    exams = q.offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for e in exams:
        chapter = db.query(Chapter).filter(Chapter.id == e.chapter_id).first()
        result.append(ExamConfigResponse(
            id=e.id,
            chapter_id=e.chapter_id,
            teacher_id=e.teacher_id,
            total_questions=e.total_questions,
            choice_count=e.choice_count,
            truefalse_count=e.truefalse_count,
            essay_count=e.essay_count,
            programming_count=e.programming_count or 0,
            knowledge_points=e.get_knowledge_points(),
            evaluation_dimensions=e.get_evaluation_dimensions(),
            created_at=e.created_at.isoformat() if e.created_at else "",
            chapter_title=chapter.title if chapter else None,
            time_limit_minutes=e.time_limit_minutes or 0,
        ))
    return {
        "items": result,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@router.delete("/exams/{exam_id}")
def delete_exam(
    exam_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """删除考核配置"""
    exam = db.query(ExamConfig).filter(
        ExamConfig.id == exam_id,
        ExamConfig.teacher_id == current_user.id
    ).first()
    if not exam:
        raise HTTPException(404, "考核不存在")

    # 删除相关考试记录和错题
    records = db.query(ExamRecord).filter(ExamRecord.exam_config_id == exam_id).all()
    for r in records:
        db.query(WrongAnswer).filter(WrongAnswer.exam_record_id == r.id).delete()
        db.delete(r)
    db.delete(exam)
    db.commit()
    return {"message": "考核已删除"}


@router.get("/exams/{exam_id}/results")
def get_exam_results(
    exam_id: int,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """查看某考核的所有学生成绩（分页）"""
    exam = db.query(ExamConfig).filter(
        ExamConfig.id == exam_id,
        ExamConfig.teacher_id == current_user.id
    ).first()
    if not exam:
        raise HTTPException(404, "考核不存在")

    q = db.query(ExamRecord).filter(
        ExamRecord.exam_config_id == exam_id,
        ExamRecord.status == "completed"
    )

    total = q.count()
    records = q.offset((page - 1) * page_size).limit(page_size).all()

    results = []
    for r in records:
        student = db.query(User).filter(User.id == r.student_id).first()
        results.append({
            "record_id": r.id,
            "student_name": student.full_name or student.username if student else "未知",
            "score": r.score,
            "dimensions_scores": r.get_dimensions_scores(),
            "completed_at": r.completed_at.isoformat() if r.completed_at else None,
        })
    return {
        "items": results,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@router.get("/exams/{exam_id}/student/{student_id}/answers")
def get_student_answers(
    exam_id: int,
    student_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """查看学生的答卷详情"""
    # 1. 验证考核属于当前教师
    exam = db.query(ExamConfig).filter(
        ExamConfig.id == exam_id,
        ExamConfig.teacher_id == current_user.id
    ).first()
    if not exam:
        raise HTTPException(404, "考核不存在")

    # 2. 验证师生关系
    rel = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.student_id == student_id,
    ).first()
    if not rel:
        raise HTTPException(403, "该学生不是你的学生")

    # 3. 查找已完成的考试记录
    record = db.query(ExamRecord).filter(
        ExamRecord.exam_config_id == exam_id,
        ExamRecord.student_id == student_id,
        ExamRecord.status == "completed"
    ).first()
    if not record:
        # 检查是否有进行中的记录
        in_progress = db.query(ExamRecord).filter(
            ExamRecord.exam_config_id == exam_id,
            ExamRecord.student_id == student_id,
            ExamRecord.status == "in_progress"
        ).first()
        if in_progress:
            raise HTTPException(404, "该学生尚未提交答卷（考核进行中）")
        raise HTTPException(404, "该学生尚未参加此考核")

    # 4. 获取学生信息和章节信息
    student = db.query(User).filter(User.id == student_id).first()
    chapter = db.query(Chapter).filter(Chapter.id == exam.chapter_id).first()

    # 5. 解析答案数据
    answers_data = record.get_answers()
    questions = answers_data.get("questions", [])
    student_answers_dict = answers_data.get("student_answers", {})

    # 6. 构建题目列表（合并题目和答案）
    question_list = []
    for i, q in enumerate(questions):
        student_ans = student_answers_dict.get(str(i), "(未作答)")
        correct_ans = q.get("answer", "")

        # 判断是否正确
        q_type = q.get("type", "")
        if not student_ans or student_ans == "(未作答)":
            is_correct = False
        elif q_type in ("choice", "judge"):
            is_correct = (str(student_ans).strip() == str(correct_ans).strip())
        else:
            # 简答题/编程题：主观题不自动判断对错
            is_correct = None

        question_list.append({
            "index": i,
            "type": q_type,
            "question": q.get("question", ""),
            "options": q.get("options"),
            "correct_answer": correct_ans,
            "student_answer": student_ans,
            "is_correct": is_correct,
        })

    return {
        "student_name": student.full_name or student.username if student else "未知",
        "chapter_title": chapter.title if chapter else "",
        "score": record.score,
        "questions": question_list,
        "dimensions_scores": record.get_dimensions_scores(),
        "report_text": record.report_text or "",
    }


# ── 学生管理 ─────────────────────────────────────

@router.get("/students")
def list_students(
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """查看自己的学生列表（分页）"""
    q = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.status == "active"
    )

    total = q.count()
    relations = q.offset((page - 1) * page_size).limit(page_size).all()

    # 获取所有考核配置总数
    total_exams = db.query(func.count(ExamConfig.id)).filter(
        ExamConfig.teacher_id == current_user.id
    ).scalar() or 0

    result = []
    for rel in relations:
        student = db.query(User).filter(User.id == rel.student_id).first()
        if not student:
            continue
        # 统计完成考核数
        completed = db.query(func.count(ExamRecord.id)).filter(
            ExamRecord.student_id == student.id,
            ExamRecord.status == "completed"
        ).scalar() or 0

        result.append(StudentInfo(
            id=student.id,
            username=student.username,
            full_name=student.full_name,
            status=rel.status,
            joined_at=rel.created_at.isoformat() if rel.created_at else "",
            completed_exams=completed,
            total_exams=total_exams,
        ))
    return {
        "items": result,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@router.get("/students/{student_id}/exams")
def get_student_exam_results(
    student_id: int,
    page: int = 1,
    page_size: int = 20,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """查看某学生的考核记录（分页）"""
    # 验证师生关系
    rel = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.student_id == student_id,
    ).first()
    if not rel:
        raise HTTPException(403, "该学生不是你的学生")

    q = db.query(ExamRecord).filter(
        ExamRecord.student_id == student_id
    ).order_by(ExamRecord.started_at.desc())

    total = q.count()
    records = q.offset((page - 1) * page_size).limit(page_size).all()

    result = []
    for r in records:
        exam_config = db.query(ExamConfig).filter(ExamConfig.id == r.exam_config_id).first()
        chapter_title = ""
        if exam_config:
            ch = db.query(Chapter).filter(Chapter.id == exam_config.chapter_id).first()
            chapter_title = ch.title if ch else ""

        result.append(StudentExamResult(
            exam_config_id=r.exam_config_id,
            chapter_title=chapter_title,
            score=r.score,
            status=r.status,
            completed_at=r.completed_at.isoformat() if r.completed_at else None,
        ))
    return {
        "items": result,
        "total": total,
        "page": page,
        "page_size": page_size,
        "total_pages": math.ceil(total / page_size) if total > 0 else 0,
    }


@router.put("/students/{student_id}")
def update_student(
    student_id: int,
    data: dict,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """编辑学生信息（用户名、姓名）"""
    rel = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.student_id == student_id,
        TeacherStudent.status == "active"
    ).first()
    if not rel:
        raise HTTPException(403, "该学生不是你的学生")

    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(404, "学生不存在")

    if data.get("username") and data["username"] != student.username:
        exists = db.query(User).filter(User.username == data["username"]).first()
        if exists:
            raise HTTPException(400, "用户名已被使用")
        student.username = data["username"]

    if "full_name" in data:
        student.full_name = data["full_name"]

    db.commit()
    return {"message": "更新成功", "username": student.username, "full_name": student.full_name}


@router.post("/students/{student_id}/reset-password")
def reset_student_password(
    student_id: int,
    data: dict,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """重置学生密码"""
    rel = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.student_id == student_id,
    ).first()
    if not rel:
        raise HTTPException(403, "该学生不是你的学生")

    student = db.query(User).filter(User.id == student_id).first()
    if not student:
        raise HTTPException(404, "学生不存在")

    new_password = data.get("new_password", "123456")
    if len(new_password) < 6:
        raise HTTPException(400, "密码至少6位")

    from ...core.security import get_password_hash
    student.password_hash = get_password_hash(new_password)
    db.commit()
    return {"message": f"密码已重置为 {new_password}"}


@router.delete("/students/{student_id}")
def remove_student(
    student_id: int,
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """移除学生（解除师生关系）"""
    rel = db.query(TeacherStudent).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.student_id == student_id,
    ).first()
    if not rel:
        raise HTTPException(404, "师生关系不存在")

    rel.status = "inactive"
    db.commit()
    return {"message": "已移除该学生"}


# ── 章节管理 ─────────────────────────────────────

@router.get("/chapters")
def get_chapters(
    db: Session = Depends(get_db)
):
    """获取所有章节列表"""
    chapters = db.query(Chapter).order_by(Chapter.order).all()
    return [
        {
            "id": c.id,
            "title": c.title,
            "description": c.description,
            "order": c.order,
            "course_name": c.course_name,
        }
        for c in chapters
    ]


# ── 模型配置 ─────────────────────────────────────

@router.get("/model-config", response_model=ModelConfigResponse)
def get_model_config(
    current_user: User = Depends(get_current_teacher),
):
    """查看 LLM 配置（API Key 脱敏）"""
    key = settings.OPENAI_API_KEY
    masked = key[:4] + "****" + key[-4:] if len(key) > 8 else "****"
    return ModelConfigResponse(
        api_key_masked=masked,
        base_url=settings.OPENAI_BASE_URL,
        model=settings.LLM_MODEL,
        embedding_model=settings.EMBEDDING_MODEL,
    )


@router.put("/model-config")
def update_model_config(
    config: ModelConfigUpdate,
    current_user: User = Depends(get_current_teacher),
):
    """更新 LLM 配置（内存生效，重启后恢复 .env 值）"""
    # 注意：此为运行时修改，不持久化到 .env 文件
    if config.api_key:
        settings.OPENAI_API_KEY = config.api_key
    if config.base_url:
        settings.OPENAI_BASE_URL = config.base_url
    if config.model:
        settings.LLM_MODEL = config.model
    if config.embedding_model:
        settings.EMBEDDING_MODEL = config.embedding_model

    return {"message": "配置已更新（运行时生效）"}


# ── 批量导入学生 ──────────────────────────────

@router.post("/students/import")
async def import_students(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """批量导入学生（支持 CSV / Excel）"""
    import io, csv
    from ...core.security import get_password_hash

    content = await file.read()
    ext = os.path.splitext(file.filename)[1].lower()
    results = {"created": [], "skipped": [], "errors": []}

    rows = []
    if ext == '.csv':
        text = content.decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(text))
        rows = list(reader)
    elif ext in ('.xlsx', '.xls'):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))
        except ImportError:
            raise HTTPException(500, "需要安装 openpyxl: pip install openpyxl")
    else:
        raise HTTPException(400, "仅支持 CSV 或 Excel (.xlsx) 文件")

    # 列名映射（支持中文/英文表头）
    col_map = {
        'username': ['username', '用户名', '账号'],
        'password': ['password', '密码'],
        'full_name': ['full_name', 'name', '姓名', '名字'],
    }

    for i, row in enumerate(rows, 2):
        try:
            # 解析列
            def get_col(key):
                for name in col_map[key]:
                    if name in row:
                        val = str(row[name]).strip()
                        return val if val and val != 'None' else None
                return None

            username = get_col('username')
            password = get_col('password') or '123456'
            full_name = get_col('full_name')

            if not username:
                results["errors"].append(f"第{i}行: 缺少用户名")
                continue
            if len(username) < 3:
                results["errors"].append(f"第{i}行: 用户名{username}太短(>=3)")
                continue

            existing = db.query(User).filter(User.username == username).first()
            if existing:
                results["skipped"].append(username)
                continue

            user = User(
                username=username,
                password_hash=get_password_hash(password),
                role="student",
                full_name=full_name
            )
            db.add(user)
            db.commit()

            # 自动关联到当前教师
            rel = TeacherStudent(teacher_id=current_user.id, student_id=user.id, status="active")
            db.add(rel)
            db.commit()

            results["created"].append(username)

        except Exception as e:
            results["errors"].append(f"第{i}行: {str(e)}")

    return {
        "message": f"导入完成: 新增{len(results['created'])}人, 跳过{len(results['skipped'])}人, 错误{len(results['errors'])}条",
        **results
    }


# ── 仪表盘 ─────────────────────────────────────

@router.get("/stats", response_model=DashboardStats)
def get_stats(
    current_user: User = Depends(get_current_teacher),
    db: Session = Depends(get_db)
):
    """教师仪表盘统计"""
    total_students = db.query(func.count(TeacherStudent.id)).filter(
        TeacherStudent.teacher_id == current_user.id,
        TeacherStudent.status == "active"
    ).scalar() or 0

    total_materials = db.query(func.count(CourseMaterial.id)).filter(
        CourseMaterial.teacher_id == current_user.id
    ).scalar() or 0

    total_exams = db.query(func.count(ExamConfig.id)).filter(
        ExamConfig.teacher_id == current_user.id
    ).scalar() or 0

    completed = db.query(func.count(ExamRecord.id)).join(
        ExamConfig, ExamRecord.exam_config_id == ExamConfig.id
    ).filter(
        ExamConfig.teacher_id == current_user.id,
        ExamRecord.status == "completed"
    ).scalar() or 0

    avg = db.query(func.avg(ExamRecord.score)).join(
        ExamConfig, ExamRecord.exam_config_id == ExamConfig.id
    ).filter(
        ExamConfig.teacher_id == current_user.id,
        ExamRecord.status == "completed"
    ).scalar() or 0

    return DashboardStats(
        total_students=total_students,
        total_materials=total_materials,
        total_exams=total_exams,
        completed_exams=completed,
        avg_score=round(float(avg), 1),
    )
